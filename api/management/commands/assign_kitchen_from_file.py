"""Bulk-assign a list of clients to a kitchen + cadence, doing EXACTLY what the
member-profile "assign kitchen" flow does for each one.

Given a file with one client id per row (default column ``MemberID``), for every
client's active enrollment this command calls
``assign_kitchen_to_household`` -- the same shared function used by the Logistics
popup and the Household tab. That means each client gets the full treatment:

  * kitchen set on the enrollment,
  * kitchen-aware meal rules applied to every member (an unfulfillable
    menu/allergy combo is flagged Out of Orbit; a now-servable member is
    reactivated),
  * the delivery plan + dated calendar (re)built with the cadence's defined
    delivery day (e.g. Rockland's Tuesday, PO cut the Friday before) -- existing
    plans have their cadence re-applied and calendar rebuilt,
  * future SCHEDULED occurrences re-synced so PO generation groups the household
    under the NEW kitchen + schedule,
  * service (re)activated (Service Active).

Dry-run by default (rolls back). Re-runnable and idempotent.

Usage:
    # DRY RUN (default file, Rockland, its only cadence):
    python manage.py assign_kitchen_from_file

    # Commit:
    python manage.py assign_kitchen_from_file --apply

    # Explicit options:
    python manage.py assign_kitchen_from_file \
        --file tmp/verification/BoxesClientsMovingtoRockland.xlsx \
        --kitchen Rockland --cadence tue_only --apply
"""
import csv
from collections import Counter

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from api.models import Client, Kitchen
from api.portal import serializers as s
from api.portal.views_members import assign_kitchen_to_household
from api.services.delivery import active_cadence_codes, cadence_needs_weekday

_DEFAULT_FILE = "tmp/verification/BoxesClientsMovingtoRockland.xlsx"


def _read_ids(path, column):
    """Read client ids from a .xlsx or .csv file. The column header is matched
    case-insensitively; a single-column file is read positionally."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            return []
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        try:
            idx = header.index(column.strip().lower())
        except ValueError:
            idx = 0  # single-column / unlabeled: use the first column
        return [str(r[idx]).strip() for r in rows[1:] if r and r[idx]]

    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        rows = list(reader)
    if not rows:
        return []
    header = [c.strip().lower() for c in rows[0]]
    idx = header.index(column.strip().lower()) if column.strip().lower() in header else 0
    return [r[idx].strip() for r in rows[1:] if r and len(r) > idx and r[idx].strip()]


class Command(BaseCommand):
    help = (
        "Bulk-assign clients from a file to a kitchen + cadence using the same "
        "assign_kitchen_to_household flow as the member profile (meal rules, "
        "schedule/calendar rebuild, PO resync, activation). Dry-run unless --apply."
    )

    def add_arguments(self, parser):
        parser.add_argument("--apply", action="store_true", help="Commit changes.")
        parser.add_argument("--file", default=_DEFAULT_FILE, help="Path to .xlsx/.csv.")
        parser.add_argument("--column", default="MemberID", help="Client-id column header.")
        parser.add_argument("--kitchen", default="Rockland", help="Target kitchen name.")
        parser.add_argument(
            "--cadence", default="",
            help="Cadence code. Defaults to the kitchen's only cadence when unambiguous.",
        )
        parser.add_argument(
            "--weekday", default="",
            help="Delivery weekday code (mon/tue/...); only for once-a-week cadences.",
        )
        parser.add_argument("--limit", type=int, default=0, help="Process first N ids.")

    def handle(self, *args, **options):
        apply = options["apply"]
        path = options["file"]

        try:
            ids = _read_ids(path, options["column"])
        except FileNotFoundError:
            raise CommandError(f"File not found: {path!r}")
        if options["limit"]:
            ids = ids[: options["limit"]]
        if not ids:
            raise CommandError("No client ids found in the file.")

        # Resolve + validate the kitchen and cadence ONCE (matches the view's
        # pre-flight checks so a misconfiguration fails fast, before any writes).
        kitchen = Kitchen.objects.filter(name__iexact=options["kitchen"].strip()).first()
        if kitchen is None:
            have = ", ".join(Kitchen.objects.values_list("name", flat=True)) or "(none)"
            raise CommandError(f"Kitchen {options['kitchen']!r} not found. Have: {have}")

        kitchen_cadences = [c.code for c in kitchen.cadences.all()]
        cadence = options["cadence"].strip()
        if not cadence:
            if len(kitchen_cadences) == 1:
                cadence = kitchen_cadences[0]
            else:
                raise CommandError(
                    f"{kitchen.name} runs {kitchen_cadences or '(none)'}; "
                    "pass --cadence to pick one."
                )
        if cadence not in active_cadence_codes():
            raise CommandError(f"Cadence {cadence!r} is not an active cadence.")
        if cadence not in kitchen_cadences:
            raise CommandError(
                f"{kitchen.name} isn't configured for cadence {cadence!r} "
                f"(runs {kitchen_cadences}). Set it in Settings -> Kitchens."
            )
        weekday = options["weekday"].strip() or None
        if cadence_needs_weekday(cadence) and not weekday:
            raise CommandError(
                f"Cadence {cadence!r} has no fixed delivery day; pass --weekday."
            )

        self.stdout.write(self.style.MIGRATE_HEADING(
            f"\nAssigning {len(ids)} client(s) -> {kitchen.name} "
            f"(cadence={cadence}{', weekday=' + weekday if weekday else ''})"
        ))

        report = Counter()
        flags = []  # (client_id, reason)

        with transaction.atomic():
            for cid in ids:
                try:
                    with transaction.atomic():
                        outcome = self._process(cid, kitchen, cadence, weekday)
                except Exception as exc:  # isolate a bad row, keep going
                    outcome = ("error", str(exc))
                report[outcome[0]] += 1
                if outcome[0] != "assigned":
                    flags.append((cid, outcome[1] if len(outcome) > 1 else ""))

            if not apply:
                transaction.set_rollback(True)

        self._report(report, flags, apply)

    def _process(self, cid, kitchen, cadence, weekday):
        client = Client.objects.filter(client_id=cid).first()
        if client is None:
            return ("skip_client_not_found", "client id not in DB")
        enr = s.active_enrollment(client)
        if enr is None:
            return ("skip_no_active_enrollment", "no active enrollment")

        assign_kitchen_to_household(
            enr, client, kitchen, cadence=cadence, once_weekday=weekday,
        )
        return ("assigned",)

    def _report(self, report, flags, apply):
        head = self.style.MIGRATE_HEADING
        self.stdout.write(head("\n=== Bulk kitchen assignment ==="))
        order = [
            ("assigned", "Assigned + schedule rebuilt + activated"),
            ("skip_no_active_enrollment", "Skipped: no active enrollment"),
            ("skip_client_not_found", "Skipped: client id not found"),
            ("error", "Errored (rolled back, see flags)"),
        ]
        for key, label in order:
            if report.get(key):
                self.stdout.write(f"  {label:<44}: {report[key]}")
        self.stdout.write(f"  {'TOTAL':<44}: {sum(report.values())}")

        if flags:
            self.stdout.write(head(f"\nFlagged ({len(flags)}, showing up to 30):"))
            for cid, reason in flags[:30]:
                self.stdout.write(f"  {cid}: {reason}")

        if apply:
            self.stdout.write(self.style.SUCCESS("\nAPPLIED (committed)."))
        else:
            self.stdout.write(self.style.WARNING(
                "\nDRY RUN: rolled back. Re-run with --apply to commit."
            ))
